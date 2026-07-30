#!/usr/bin/env python3
"""
name_sweep.py — deterministic name search across a conflicts-check workbook.

Usage:
    python name_sweep.py <workbook.xlsx> "Name One" "Name Two" ...

Sweeps EVERY cell of EVERY tab (values only), so names buried in Notes columns
(e.g., end-client counterparties recorded in Parties Index notes) are caught.

Matching:
  - case-insensitive
  - punctuation-normalized
  - entity suffixes stripped (Inc, LLC, Corp, Co, Ltd, LP, LLP, PLLC, Company,
    Corporation, Incorporated, GmbH, SA, NV)
  - substring match in BOTH directions after normalization ("acme" hits
    "Acme Robotics Corporation"). Note that an acronym or shorthand will only
    hit if an alias row exists in the workbook — which is why alias rows matter.
  - token-overlap fallback: if >=2 significant tokens (len>3) of the query appear
    in a cell, report as PARTIAL

Output: JSON to stdout — one entry per hit with tab, cell, row context (full row
values), match type.

Exit code 0 always; an empty "hits" list means no matches. That is not legal
clearance — the analysis is the lawyer's job.
"""
import sys, json, re
import openpyxl

SUFFIXES = r"\b(incorporated|corporation|company|limited|inc|llc|corp|co|ltd|lp|llp|pllc|plc|gmbh|sa|nv)\b\.?"

def normalize(s):
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(SUFFIXES, " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def tokens(s):
    return {t for t in normalize(s).split() if len(t) > 3}

def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "usage: name_sweep.py <workbook.xlsx> NAME [NAME ...]"}))
        return
    path, names = sys.argv[1], sys.argv[2:]
    wb = openpyxl.load_workbook(path, data_only=True)
    results = {}
    for name in names:
        qn = normalize(name)
        qt = tokens(name)
        hits = []
        for ws in wb.worksheets:
            headers = [str(c.value) if c.value else "" for c in ws[1]]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    cn = normalize(cell.value)
                    if not cn or not qn:
                        continue
                    match = None
                    if qn == cn:
                        match = "EXACT"
                    elif qn in cn or cn in qn:
                        match = "SUBSTRING"
                    elif qt and len(qt & tokens(cell.value)) >= 2:
                        match = "PARTIAL"
                    if match:
                        ctx = {h: (str(c.value) if c.value is not None else "")
                               for h, c in zip(headers, row) if c.value is not None}
                        hits.append({"tab": ws.title, "cell": cell.coordinate,
                                     "match": match, "matched_text": cell.value[:120],
                                     "row_context": ctx})
                        break  # one hit per row per name is enough
        results[name] = hits
    print(json.dumps({"workbook": path, "results": results}, indent=1))

if __name__ == "__main__":
    main()
